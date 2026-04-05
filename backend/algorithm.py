import pandas as pd
import math
import itertools
from collections import defaultdict, Counter
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
from io import BytesIO


DEFAULT_MAX_CLASSES = 20
DEFAULT_ROWS = 10
DEFAULT_COLS = 4
DEFAULT_MIN_PERFECT = 30
DEFAULT_MIN_GOOD = 20
DEFAULT_RANDOM_SEED = 42


def find_col(df, candidates):
    for c in candidates:
        if c in df.columns:
            return c
    for c in df.columns:
        if df[c].dtype == object:
            return c
    return df.columns[0]


def create_ab_room(d_A, d_B, remaining_counts, student_ptrs, dept_lists,
                   min_students, room_id, rows, cols, notes_prefix):
    """
    Fill all `cols` columns using a strict alternating [A, B, A, B, A, ...]
    pattern across however many columns the room has.
    - Odd-indexed columns (0, 2, 4, ...) → Dept A
    - Even-indexed columns (1, 3, 5, ...) → Dept B
    This works correctly for any number of columns (4, 5, 6, etc.)
    """
    # Count how many columns are assigned to each dept
    cols_A = [c for c in range(cols) if c % 2 == 0]   # cols 0,2,4,...
    cols_B = [c for c in range(cols) if c % 2 == 1]   # cols 1,3,5,...

    # How many students can fill each dept's columns (rows per column)
    avail_A = remaining_counts.get(d_A, 0)
    avail_B = remaining_counts.get(d_B, 0)

    # For each dept, distribute students across their assigned columns
    # Each column holds at most `rows` students
    def plan_fill(avail, num_cols):
        """Return list of how many students go in each of this dept's columns."""
        fills = []
        remaining = avail
        for _ in range(num_cols):
            take = min(remaining, rows)
            fills.append(take)
            remaining -= take
            if remaining <= 0:
                break
        # Pad with zeros if dept runs out before filling all columns
        while len(fills) < num_cols:
            fills.append(0)
        return fills

    fills_A = plan_fill(avail_A, len(cols_A))
    fills_B = plan_fill(avail_B, len(cols_B))

    total = sum(fills_A) + sum(fills_B)
    if total < min_students:
        return None, None

    seats_per_room = rows * cols
    seats = [None] * seats_per_room
    students_taken_A = 0
    students_taken_B = 0

    # Fill Dept A columns
    for i, col_idx in enumerate(cols_A):
        for r in range(fills_A[i]):
            seat_idx = r * cols + col_idx
            if seat_idx < seats_per_room and student_ptrs[d_A] < len(dept_lists[d_A]):
                seats[seat_idx] = dept_lists[d_A][student_ptrs[d_A]]
                student_ptrs[d_A] += 1
                students_taken_A += 1

    # Fill Dept B columns
    for i, col_idx in enumerate(cols_B):
        for r in range(fills_B[i]):
            seat_idx = r * cols + col_idx
            if seat_idx < seats_per_room and student_ptrs[d_B] < len(dept_lists[d_B]):
                seats[seat_idx] = dept_lists[d_B][student_ptrs[d_B]]
                student_ptrs[d_B] += 1
                students_taken_B += 1

    remaining_counts[d_A] = remaining_counts.get(d_A, 0) - students_taken_A
    remaining_counts[d_B] = remaining_counts.get(d_B, 0) - students_taken_B

    room_data = {"room_no": room_id, "students": seats, "departments": {d_A, d_B}}
    room_note = f"OK ({notes_prefix})"
    return room_data, room_note


def compute_metrics(rooms, room_notes, room_invigilators, teachers_df, rows, cols):
    """Compute fairness index, utilization, load variance, conflicts."""
    seats_per_room = rows * cols
    total_seats = len(rooms) * seats_per_room
    filled_seats = sum(
        sum(1 for s in room["students"] if s is not None) for room in rooms
    )
    utilization = round((filled_seats / total_seats * 100) if total_seats > 0 else 0, 2)

    # Fairness index: how evenly students are distributed across departments
    dept_counts = defaultdict(int)
    for room in rooms:
        for s in room["students"]:
            if s:
                dept_counts[s["Department"]] += 1
    if dept_counts:
        counts = list(dept_counts.values())
        mean_count = sum(counts) / len(counts)
        variance = sum((c - mean_count) ** 2 for c in counts) / len(counts)
        max_variance = mean_count ** 2 if mean_count > 0 else 1
        fairness_index = round(1 - (variance / max_variance) if max_variance > 0 else 1, 2)
        fairness_index = max(0.0, min(1.0, fairness_index))
    else:
        fairness_index = 0.0

    # Load variance: std dev of invigilator duties
    load_counter = Counter()
    conflict_count = 0
    for room in rooms:
        invigs = room_invigilators.get(room["room_no"], [])
        for inv in invigs:
            if inv["TeacherID"] != "N/A":
                load_counter[inv["TeacherID"]] += 1
            # Check conflict: invigilator dept in room depts
            if inv["Department"] in room["departments"]:
                conflict_count += 1

    loads = list(load_counter.values())
    if loads:
        mean_load = sum(loads) / len(loads)
        load_variance = round(
            math.sqrt(sum((l - mean_load) ** 2 for l in loads) / len(loads)), 2
        )
    else:
        load_variance = 0.0

    return {
        "room_utilization": utilization,
        "fairness_index": fairness_index,
        "load_variance": load_variance,
        "conflict_count": conflict_count,
        "total_rooms": len(rooms),
        "total_students": filled_seats,
        "dept_distribution": dict(dept_counts),
    }


def generate_seating_plan(students_df, teachers_df, max_classes=DEFAULT_MAX_CLASSES,
                           rows=DEFAULT_ROWS, cols=DEFAULT_COLS,
                           min_perfect=DEFAULT_MIN_PERFECT, min_good=DEFAULT_MIN_GOOD,
                           random_seed=DEFAULT_RANDOM_SEED, shuffle_within_dept=False,
                           ai_constraints=None):
    logs = []
    unallocated_log = []

    # Normalise AI constraints — default to empty if not provided
    if ai_constraints is None:
        ai_constraints = {}
    forbidden_pairs = [
        (str(p[0]), str(p[1]))
        for p in ai_constraints.get("forbidden_dept_pairs", [])
    ]
    if forbidden_pairs:
        logs.append(f"[AI] Forbidden dept pairs: {forbidden_pairs}")
    unavailable_log = ai_constraints.get("unavailable_teachers", [])
    if unavailable_log:
        logs.append(f"[AI] Unavailable teachers: {unavailable_log}")
    fixed_log = ai_constraints.get("fixed_invigilators", [])
    if fixed_log:
        logs.append(f"[AI] Fixed invigilator assignments: {fixed_log}")

    students = students_df.copy()
    teachers = teachers_df.copy()
    logs.append(f"Loaded {len(students)} students and {len(teachers)} teachers.")

    student_name_col = find_col(students, ["Name", "StudentName", "FullName", "name", "student_name", "Student Name"])
    student_id_col = find_col(students, ["StudentID", "ID", "Roll", "RollNo", "RegNo", "RegistrationNo", "roll"])
    student_dept_col = find_col(students, ["Department", "Dept", "department", "dept", "Department Name"])

    teacher_name_col = find_col(teachers, ["Name", "TeacherName", "FullName", "name", "teacher_name", "Teacher Name"])
    teacher_id_col = find_col(teachers, ["TeacherID", "ID", "Tid", "teacher_id"])
    teacher_dept_col = find_col(teachers, ["Department", "Dept", "department", "dept", "Department Name"])

    students = students.rename(columns={
        student_name_col: "Name", student_id_col: "StudentID", student_dept_col: "Department"
    })
    teachers = teachers.rename(columns={
        teacher_name_col: "Name", teacher_id_col: "TeacherID", teacher_dept_col: "Department"
    })

    students = students[["StudentID", "Name", "Department"]].copy()
    teachers = teachers[["TeacherID", "Name", "Department"]].copy()

    seats_per_room = rows * cols
    capacity = max_classes * seats_per_room
    if len(students) > capacity:
        logs.append(f"More students ({len(students)}) than capacity ({capacity}). Truncating.")
        students = students.sample(n=capacity, random_state=random_seed).reset_index(drop=True)

    dept_students = {}
    for d, df in students.groupby("Department"):
        if shuffle_within_dept:
            df = df.sample(frac=1, random_state=random_seed).reset_index(drop=True)
        else:
            df = df.reset_index(drop=True)
        dept_students[d] = df

    departments = list(dept_students.keys())
    dept_lists = {d: list(df.to_dict("records")) for d, df in dept_students.items()}
    student_ptrs = defaultdict(int)
    remaining = {d: len(dept_students[d]) for d in departments}

    rooms = []
    room_notes = {}
    room_id = 1

    def is_forbidden(d1, d2):
        """Check if this dept pair is forbidden by AI constraints."""
        return any(
            (d1 == p[0] and d2 == p[1]) or (d1 == p[1] and d2 == p[0])
            for p in forbidden_pairs
        )

    def estimate_ab_capacity(avail_A, avail_B, rows, cols):
        cols_A = (cols + 1) // 2
        cols_B = cols // 2
        return min(avail_A, cols_A * rows) + min(avail_B, cols_B * rows)

    # Phase 1: Perfect rooms
    logs.append(f"--- Phase 1: 'Perfect' rooms (Min {min_perfect}) ---")
    while room_id <= max_classes:
        available_depts = [d for d, c in remaining.items() if c > 0]
        if len(available_depts) < 2:
            break
        best_pair, max_students = None, 0
        for d1, d2 in itertools.permutations(available_depts, 2):
            if is_forbidden(d1, d2):
                continue
            total = estimate_ab_capacity(remaining[d1], remaining[d2], rows, cols)
            if total >= min_perfect and total > max_students:
                max_students = total; best_pair = (d1, d2)
        if not best_pair:
            break
        d_A, d_B = best_pair
        room_data, room_note = create_ab_room(d_A, d_B, remaining, student_ptrs, dept_lists,
                                               min_perfect, room_id, rows, cols, f"[A,B,...] Layout, >={min_perfect}")
        if room_data:
            rooms.append(room_data); room_notes[room_id] = room_note; room_id += 1
        else:
            break

    # Phase 2: Good rooms
    logs.append(f"--- Phase 2: 'Good' rooms (Min {min_good}) ---")
    while room_id <= max_classes:
        available_depts = [d for d, c in remaining.items() if c > 0]
        if len(available_depts) < 2:
            break
        best_pair, max_students = None, 0
        for d1, d2 in itertools.permutations(available_depts, 2):
            if is_forbidden(d1, d2):
                continue
            total = estimate_ab_capacity(remaining[d1], remaining[d2], rows, cols)
            if total >= min_good and total > max_students:
                max_students = total; best_pair = (d1, d2)
        if not best_pair:
            break
        d_A, d_B = best_pair
        room_data, room_note = create_ab_room(d_A, d_B, remaining, student_ptrs, dept_lists,
                                               min_good, room_id, rows, cols, f"[A,B,...] Layout, >={min_good}")
        if room_data:
            rooms.append(room_data); room_notes[room_id] = room_note; room_id += 1
        else:
            break

    # Phase 3: Leftover columnar
    logs.append("--- Phase 3: Leftover columnar rooms ---")
    while room_id <= max_classes and any(v > 0 for v in remaining.values()):
        depts_sorted = sorted([(d, c) for d, c in remaining.items() if c > 0], key=lambda x: -x[1])
        if not depts_sorted:
            break
        col_depts = [d[0] for d in depts_sorted[:cols]]
        room_depts_set = set()
        seats = [None] * (rows * cols)
        col_students = {c: [] for c in range(cols)}
        students_in_room = 0
        for c, dept in enumerate(col_depts):
            take = min(remaining.get(dept, 0), rows)
            if take > 0:
                room_depts_set.add(dept)
            for _ in range(take):
                if student_ptrs[dept] < len(dept_lists[dept]):
                    col_students[c].append(dept_lists[dept][student_ptrs[dept]])
                    student_ptrs[dept] += 1
                    remaining[dept] -= 1
                    students_in_room += 1
        for r in range(rows):
            for c in range(cols):
                if r < len(col_students[c]):
                    seats[r * cols + c] = col_students[c][r]
        if students_in_room > 0:
            rooms.append({"room_no": room_id, "students": seats, "departments": room_depts_set})
            room_notes[room_id] = ("OK (Columnar Leftover, >2 Depts)" if len(room_depts_set) > 2
                                   else "OK (Columnar Leftover)")
            room_id += 1

    unallocated_count = sum(remaining.values())
    if unallocated_count > 0:
        unallocated_log.append(f"WARNING: {unallocated_count} students remain unallocated.")
        for d, c in remaining.items():
            if c > 0:
                unallocated_log.append(f"   - {d}: {c} students")
    else:
        logs.append("All students successfully allocated.")

    # ── Invigilator assignment (AI-constraint-aware, double-booking-safe) ──
    # Normalize AI constraints
    unavailable_ids  = [str(x) for x in ai_constraints.get("unavailable_teachers", [])]
    fixed_assignments = ai_constraints.get("fixed_invigilators", [])
    pre_booked_ids   = [str(f.get("TeacherID")) for f in fixed_assignments]

    teacher_pool = teachers.to_dict("records")
    # Force TeacherID to string so it matches AI JSON output
    for t in teacher_pool:
        t["TeacherID"] = str(t["TeacherID"])

    teacher_load       = Counter()
    room_invigilators  = {}
    # Track globally assigned teachers to prevent double-booking
    assigned_in_session = set(pre_booked_ids)

    for room in rooms:
        room_no      = room["room_no"]
        forbidden    = room["departments"]
        chosen       = []

        # 1. Apply AI fixed-room assignments first
        for fixed in fixed_assignments:
            if fixed.get("Room") == room_no:
                fixed_tid = str(fixed.get("TeacherID"))
                teacher   = next((t for t in teacher_pool if t["TeacherID"] == fixed_tid), None)
                if teacher and fixed_tid not in unavailable_ids:
                    chosen.append(teacher)
                    teacher_load[fixed_tid] += 1
                    logs.append(f"[AI] Fixed assignment: {teacher['Name']} → Room {room_no}")

        # 2. Build valid candidate pool:
        #    - Not unavailable (AI leave constraint)
        #    - Not already assigned to another room (double-booking prevention)
        #    - Not from a forbidden department (same-dept constraint)
        valid = [
            t for t in teacher_pool
            if t["TeacherID"] not in unavailable_ids
            and t["TeacherID"] not in assigned_in_session
            and t["Department"] not in forbidden
        ]
        valid = sorted(valid, key=lambda x: teacher_load[x["TeacherID"]])

        # 3. Fill up to 2 invigilators
        while len(chosen) < 2:
            selected = None

            if valid:
                # For the 2nd slot, try to pick a different department than 1st
                if len(chosen) == 1:
                    chosen_dept = chosen[0]["Department"]
                    for i, c in enumerate(valid):
                        if c["Department"] != chosen_dept:
                            selected = valid.pop(i)
                            break
                # Fallback: just take lowest-load valid teacher
                if not selected:
                    selected = valid.pop(0)
            else:
                # Fallback: relax same-dept constraint but still prevent double-booking
                fallback = [
                    t for t in teacher_pool
                    if t["TeacherID"] not in unavailable_ids
                    and t["TeacherID"] not in assigned_in_session
                ]
                fallback = sorted(fallback, key=lambda x: teacher_load[x["TeacherID"]])
                if fallback:
                    selected = fallback.pop(0)
                    logs.append(f"[Warning] Relaxed dept constraint for Room {room_no}: assigned {selected['Name']}")

            if selected:
                chosen.append(selected)
                assigned_in_session.add(selected["TeacherID"])
                teacher_load[selected["TeacherID"]] += 1
            else:
                chosen.append({"Name": "N/A", "TeacherID": "N/A", "Department": "N/A"})
                break

        room_invigilators[room_no] = chosen

    logs.append("Invigilator assignment complete.")

    # Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Seating Plan"
    fill_colors = [
        PatternFill(start_color="FFFFE0", end_color="FFFFE0", fill_type="solid"),
        PatternFill(start_color="E0FFFF", end_color="E0FFFF", fill_type="solid"),
        PatternFill(start_color="F0FFF0", end_color="F0FFF0", fill_type="solid"),
        PatternFill(start_color="FFF0F5", end_color="FFF0F5", fill_type="solid"),
    ]
    good_fill = PatternFill(start_color="E0FFE0", end_color="E0FFE0", fill_type="solid")
    warning_fill = PatternFill(start_color="FFFACD", end_color="FFFACD", fill_type="solid")

    start_row = 1
    for i, room in enumerate(rooms):
        r = start_row
        room_depts_str = ", ".join(sorted(room["departments"]))
        note = room_notes.get(room["room_no"], "OK")

        if ">2 Depts" in note or "Leftover" in note:
            room_title = f"Room {room['room_no']} (Depts: {room_depts_str}) - LEFTOVER ROOM"
            cell_fill = warning_fill
        else:
            room_title = f"Room {room['room_no']} (Depts: {room_depts_str}) - Layout: [A, B, A, B]"
            cell_fill = fill_colors[i % len(fill_colors)] if "perfect" not in note.lower() else good_fill

        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=cols)
        cell = ws.cell(row=r, column=1, value=room_title)
        cell.fill = cell_fill
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
        r += 1

        for c in range(cols):
            col_cell = ws.cell(row=r, column=c+1, value=f"Column {c+1}")
            col_cell.font = Font(bold=True)
        r += 1

        for rowi in range(rows):
            ws.row_dimensions[r + rowi].height = 40
            for coli in range(cols):
                idx = rowi * cols + coli
                cell = ws.cell(row=r+rowi, column=coli+1)
                if idx < len(room["students"]) and room["students"][idx] is not None:
                    s = room["students"][idx]
                    cell.value = f"{s['Name']}\n({s['StudentID']})\n{s['Department']}"
                else:
                    cell.value = "--- EMPTY ---"
                    cell.font = Font(color="999999")
                cell.alignment = Alignment(wrap_text=True, horizontal="center", vertical="center")
        r += rows

        inv_row = r + 1
        ws.merge_cells(start_row=inv_row, start_column=1, end_row=inv_row, end_column=cols)
        inv_text = "Invigilators: " + ", ".join(
            [f"{t['Name']} ({t['Department']})" for t in room_invigilators[room["room_no"]]]
        )
        inv_cell = ws.cell(row=inv_row, column=1, value=inv_text)
        inv_cell.font = Font(italic=True)
        inv_cell.alignment = Alignment(horizontal="center")
        start_row = inv_row + 3

    for col in range(1, cols + 1):
        ws.column_dimensions[get_column_letter(col)].width = 30

    # Summaries
    summary_rows = []
    for room in rooms:
        total_students = sum(1 for s in room["students"] if s is not None)
        summary_rows.append({
            "Room": room["room_no"],
            "Departments": ", ".join(sorted(room["departments"])),
            "Students_Assigned": total_students,
            "Invigilators": "; ".join([f"{t['Name']} ({t['Department']})"
                                        for t in room_invigilators[room["room_no"]]]),
            "Constraint_Check": room_notes.get(room["room_no"], "OK"),
        })
    summary_df = pd.DataFrame(summary_rows)

    load_summary_data = []
    for teacher_id, count in teacher_load.most_common():
        if teacher_id == "N/A":
            load_summary_data.append({"Teacher": "N/A", "Duties": count})
            continue
        t_name_series = teachers[teachers["TeacherID"] == teacher_id]["Name"]
        name = t_name_series.values[0] if not t_name_series.empty else f"ID:{teacher_id}"
        load_summary_data.append({"Teacher": name, "Duties": count})
    teacher_load_df = pd.DataFrame(load_summary_data) if load_summary_data else pd.DataFrame({"Teacher": ["None"], "Duties": [0]})

    # Compute metrics
    metrics = compute_metrics(rooms, room_notes, room_invigilators, teachers, rows, cols)

    excel_io = BytesIO()
    wb.save(excel_io)
    excel_io.seek(0)

    return {
        "excel_bytes": excel_io.getvalue(),
        "summary": summary_df.to_dict("records"),
        "teacher_load": teacher_load_df.to_dict("records"),
        "unallocated_log": unallocated_log,
        "logs": logs,
        "metrics": metrics,
    }
